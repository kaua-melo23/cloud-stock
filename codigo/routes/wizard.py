import os
import uuid
import re
import pytz
from datetime import datetime
from flask import Blueprint, request, jsonify
from flask_login import current_user, login_required
from openpyxl import load_workbook
from database import db, audit_log
from auth import is_analyst_or_admin
from config import gsheets, settings

bp = Blueprint('wizard_sys', __name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads_temp')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def get_now_br():
    return datetime.now(pytz.timezone('America/Recife')).strftime('%Y-%m-%d %H:%M:%S')

# ... (Rotas wizard_analyze e wizard_headers mantidas iguais) ...
@bp.route('/api/wizard/analyze', methods=['POST'])
@is_analyst_or_admin
def wizard_analyze():
    if 'file' in request.files:
        f = request.files['file']
        fid = f"{uuid.uuid4()}_{f.filename}"
        path = os.path.join(UPLOAD_FOLDER, fid)
        f.save(path)
        try: return jsonify({'success': True, 'mode': 'file', 'id': fid, 'sheets': load_workbook(path, read_only=True).sheetnames})
        except Exception as e: return jsonify({'message': f"Erro arquivo: {str(e)}"}), 500
    d = request.get_json() or {}
    if 'url' in d:
        sid = (re.search(r"/d/([a-zA-Z0-9-_]+)", d['url']) or d).group(1) if re.search(r"/d/([a-zA-Z0-9-_]+)", d['url']) else d['url']
        try: return jsonify({'success': True, 'mode': 'link', 'id': sid, 'sheets': [s['properties']['title'] for s in gsheets.get_sheet_metadata(sid).get('sheets', [])]})
        except Exception as e: return jsonify({'message': f"Erro Google: {str(e)}"}), 500
    return jsonify({'message': 'Inválido'}), 400

@bp.route('/api/wizard/headers', methods=['POST'])
@is_analyst_or_admin
def wizard_headers():
    d = request.get_json()
    try:
        rows = []
        if d['mode'] == 'file':
            wb = load_workbook(os.path.join(UPLOAD_FOLDER, d['id']), read_only=True)
            rows = [next(wb[d['sheet_name']].iter_rows(values_only=True))]
            wb.close()
        else: rows = gsheets.read_sheet(f"{d['sheet_name']}!A1:Z1", spreadsheet_id=d['id'])
        return jsonify({'headers': [str(h).strip() for h in rows[0]] if rows else []})
    except Exception as e: return jsonify({'message': str(e)}), 500

# --- CORREÇÃO DO ERRO 500 AQUI ---
@bp.route('/api/wizard/unique-values', methods=['POST'])
@is_analyst_or_admin
def wizard_unique():
    d = request.get_json()
    try:
        vals = set()
        # Converte para int e protege contra erro
        try:
            col_idx = int(d['column_index'])
        except:
            return jsonify({'message': 'Índice de coluna inválido'}), 400

        if d['mode'] == 'file':
            wb = load_workbook(os.path.join(UPLOAD_FOLDER, d['id']), read_only=True)
            for r in wb[d['sheet_name']].iter_rows(min_row=2, values_only=True): 
                if col_idx < len(r) and r[col_idx]: 
                    vals.add(str(r[col_idx]).strip())
            wb.close()
        else:
            # Google Sheets
            data = gsheets.read_sheet(d['sheet_name'], spreadsheet_id=d['id'])
            # Pula cabeçalho (start=1)
            for r in data[1:]:
                if col_idx < len(r) and r[col_idx]: 
                    vals.add(str(r[col_idx]).strip())
                    
        return jsonify({'values': sorted(list(vals))})
    except Exception as e: 
        print(f"Erro Unique Values: {e}") # Log no console do servidor
        return jsonify({'message': f"Erro interno: {str(e)}"}), 500

@bp.route('/api/wizard/execute', methods=['POST'])
@is_analyst_or_admin
def wizard_execute():
    d = request.get_json()
    try:
        rows = []
        if d['mode'] == 'file':
            wb = load_workbook(os.path.join(UPLOAD_FOLDER, d['id']), read_only=True)
            rows = list(wb[d['sheet_name']].iter_rows(min_row=2, values_only=True))
            wb.close()
        else: rows = gsheets.read_sheet(d['sheet_name'], spreadsheet_id=d['id'])[1:]
        
        conn, ins, upd, now = db.get_connection('inventory'), 0, 0, get_now_br()
        cmap, vmap, ftype = d.get('column_map',{}), d.get('value_map',{}), d.get('fixed_type')
        
        for r in rows:
            if not r: continue
            def get_val(k): 
                try: 
                    idx = int(cmap.get(k))
                    return str(r[idx]).strip() if idx<len(r) and r[idx] else ""
                except: return ""

            sn = get_val('serial_number')
            if not sn or sn.lower()=='none': continue
            
            st = vmap.get('status',{}).get(get_val('status'), 'Available')
            if st=='IGNORE': continue
            
            typ = get_val('type') or (ftype.upper() if ftype else "OUTRO")

            if conn.execute('SELECT 1 FROM inventory WHERE serial_number=%s',(sn,)).fetchone():
                conn.execute('UPDATE inventory SET type=%s, model=%s, tag=%s, tombamento=%s, status=%s, glpi_status=%s, standardized=%s, last_updated=%s WHERE serial_number=%s', 
                             (typ, get_val('model') or "Genérico", get_val('tag'), get_val('tombamento'), st, vmap.get('glpi_status',{}).get(get_val('glpi_status'),'Pendente'), vmap.get('standardized',{}).get(get_val('standardized'),'Não'), now, sn))
                upd+=1
            else:
                conn.execute('INSERT INTO inventory (serial_number,type,model,tag,tombamento,condition,status,glpi_status,standardized,last_updated) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                             (sn, typ, get_val('model') or "Genérico", get_val('tag'), get_val('tombamento'), 'Novo', st, vmap.get('glpi_status',{}).get(get_val('glpi_status'),'Pendente'), vmap.get('standardized',{}).get(get_val('standardized'),'Não'), now))
                ins+=1
        conn.commit()
        audit_log(current_user.id, f"Wizard {d['mode']}", f"I:{ins}, U:{upd}")
        return jsonify({'success': True, 'message': f"Ok: {ins} novos, {upd} atualizados."})
    except Exception as e: return jsonify({'message': str(e)}), 500